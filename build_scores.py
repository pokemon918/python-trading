import pandas as pd
from datetime import datetime, timedelta
import openpyxl

from database_builds import get_build_scores
from database_strategies import get_all_trades
from market_reference import market_contract_size
from strategy_scoring import create_requirements_limits
import settings

def calculate_finder_and_improver_stats(build_scores):
    """
    Calculate stats for Finder (improver_id == 0) and Improver (improver_id != 0) items.
    Include both cumulative totals and metrics for the past 24 hours.
    """
    # Split data into Finder and Improver subsets
    finder_scores = build_scores[build_scores['improver_id'] == 0]
    improver_scores = build_scores[build_scores['improver_id'] != 0]

    # Function to calculate stats for a given subset
    def calculate_subset_stats(subset, name):
        # Explicitly create a copy to avoid the SettingWithCopyWarning
        subset = subset.copy()

        # Convert creation_datetime to datetime
        subset['creation_datetime'] = pd.to_datetime(subset['creation_datetime'])

        last_24_hours = datetime.now() - timedelta(days=1)
        recent_subset = subset[subset['creation_datetime'] >= last_24_hours]

        # Total Return per Market
        total_return_per_market = subset.groupby('market')['forward_week_returns'].sum().reset_index()
        total_return_per_market.rename(columns={'forward_week_returns': 'total_return'}, inplace=True)

        # Return Added in the Last 24 Hours
        return_added_per_market = recent_subset.groupby('market')['forward_week_returns'].sum().reset_index()
        return_added_per_market.rename(columns={'forward_week_returns': 'return_added_24h'}, inplace=True)

        # Unique Strategy IDs and Rows Added in the Last 24 Hours
        unique_ids_and_rows_per_market_24h = recent_subset.groupby('market').agg(
            unique_strategy_ids_24h=('strategy_id', 'nunique'),
            rows_added_24h=('strategy_id', 'size')
        ).reset_index()

        # Total Unique Strategy IDs and Total Rows
        total_unique_ids_and_rows = subset.groupby('market').agg(
            total_unique_strategy_ids=('strategy_id', 'nunique'),
            total_rows=('strategy_id', 'size')
        ).reset_index()

        # Merge all metrics into a single table
        stats = pd.merge(total_return_per_market, total_unique_ids_and_rows, on='market', how='outer')
        stats = pd.merge(stats, unique_ids_and_rows_per_market_24h, on='market', how='outer')
        stats = pd.merge(stats, return_added_per_market, on='market', how='outer')

        # Add Totals Row
        totals_row = {
            'market': 'Total',
            'total_return': total_return_per_market['total_return'].sum(),
            'total_unique_strategy_ids': subset['strategy_id'].nunique(),
            'total_rows': subset.shape[0],
            'unique_strategy_ids_24h': recent_subset['strategy_id'].nunique(),
            'rows_added_24h': recent_subset.shape[0],
            'return_added_24h': return_added_per_market['return_added_24h'].sum()
        }
        stats = pd.concat([stats, pd.DataFrame([totals_row])], ignore_index=True)

        return stats

    # Calculate stats for both subsets
    finder_stats = calculate_subset_stats(finder_scores, "Finder Statistics")
    improver_stats = calculate_subset_stats(improver_scores, "Improver Statistics")

    return finder_stats, improver_stats

def calculate_usd_returns(build_scores, trades):

    # Insert the usd column after the existing returns column
    forward_week_returns_idx = build_scores.columns.get_loc('forward_week_returns')
    build_scores.insert(forward_week_returns_idx + 1, 'forward_week_returns_usd', 0.0)

    for index, row in build_scores.iterrows():
        strategy_id = row['strategy_id']
        optimisation_date = row['optimisation_date']
        market = row['market']
                     
        forward_week_end_date = optimisation_date + pd.Timedelta(weeks=1)
        forward_trades = trades[(trades['strategy_id'] == strategy_id) & 
                 (trades['entry_datetime'] >= optimisation_date) & 
                 (trades['entry_datetime'] <= forward_week_end_date)]
        
        contract_size = market_contract_size[market]
        
        total_usd = 0.0
        for trade_index, trade_row in forward_trades.iterrows():
            entry_price = trade_row['entry_price']
            trade_return = trade_row['return']

            total_usd += trade_return * entry_price * contract_size
        
        build_scores.at[index, 'forward_week_returns_usd'] = total_usd

# Out of sample edge per strategy based on in sample TAWL
def calculate_oos_edge_statistics(build_scores):
    grouped = build_scores.groupby('strategy_id').agg(
        market=('market', 'first'),
        average_forward_week_returns=('forward_week_returns', 'mean'),
        weeks_in_buildscores=('strategy_id', 'size'),
        sum_forward_week_trade_count=('forward_week_trade_count', 'sum'),
        average_score=('score', 'mean'),
        average_edge_better_than_random0=('edge_better_than_random0', 'mean'),
        sum_forward_week_trade_win_count=('forward_week_trade_win_count', 'sum'),
        average_trade_win_over_loss0=('trade_win_over_loss0', 'mean')
    ).reset_index()

    # Calculate additional columns
    grouped['oos_winrate'] = grouped['sum_forward_week_trade_win_count'] / grouped['sum_forward_week_trade_count']
    grouped['random_winrate'] = 1 - grouped['average_trade_win_over_loss0'] / (grouped['average_trade_win_over_loss0'] + 1)
    grouped['oos_edge'] = grouped['oos_winrate'] - grouped['random_winrate']

    # Drop intermediate calculation column
    grouped.drop(columns=['sum_forward_week_trade_win_count'], inplace=True)

    # Add a new column for each unique market in build_scores
    unique_markets = build_scores['market'].unique()
    for market in unique_markets:
        grouped[market] = grouped.apply(
            lambda row: row['oos_edge'] if row['market'] == market else '#N/A',
            axis=1
        )

    return grouped

def calculate_improvements_statistics(build_scores):
    """
    Extracts unique strategy IDs with non-zero improver IDs and calculates various statistics.
    Returns a DataFrame with columns: "Child", "Parent", "Child Return", "Parent Return",
    "Change", "Total Child Return", "Total Parent Return after Child Optimisation Date",
    "Child Weeks", and "Parent Weeks", sorted by "Child".
    """
    # Filter rows where improver_id is non-zero
    non_zero_improvers = build_scores[build_scores['improver_id'] != 0]

    # Create the base DataFrame with Child and Parent columns
    improvements = non_zero_improvers[['strategy_id', 'improver_id']].drop_duplicates()
    improvements.rename(columns={'strategy_id': 'Child', 'improver_id': 'Parent'}, inplace=True)

    # Add the "Market" column based on the Child strategy_id
    improvements['Market'] = improvements['Child'].apply(
        lambda child: build_scores[build_scores['strategy_id'] == child]['market'].iloc[0] if not build_scores[
            build_scores['strategy_id'] == child].empty else None
    )

    # Function to calculate the returns for a given Child or Parent
    def calculate_returns(row, child_or_parent):
        # Filter rows for the Child strategy
        child_data = build_scores[build_scores['strategy_id'] == row['Child']]
        # Filter rows for the Parent strategy
        parent_data = build_scores[build_scores['strategy_id'] == row['Parent']]

        # Find common optimisation dates between Child and Parent
        common_dates = set(child_data['optimisation_date']).intersection(parent_data['optimisation_date'])

        # Calculate the sum of forward_week_returns for the specified strategy and common dates
        if child_or_parent == "Child":
            return child_data[child_data['optimisation_date'].isin(common_dates)]['forward_week_returns'].sum()
        elif child_or_parent == "Parent":
            return parent_data[parent_data['optimisation_date'].isin(common_dates)]['forward_week_returns'].sum()

    # Calculate "Child Return" for each row
    improvements['Child Return'] = improvements.apply(calculate_returns, axis=1, child_or_parent="Child")

    # Calculate "Parent Return" for each row
    improvements['Parent Return'] = improvements.apply(calculate_returns, axis=1, child_or_parent="Parent")

    # Calculate "Change" as Child Return - Parent Return
    improvements['Change'] = improvements['Child Return'] - improvements['Parent Return']

    # Calculate "Total Child Return" as the sum of all forward_week_returns for the Child strategy
    def calculate_total_child_return(row):
        child_data = build_scores[build_scores['strategy_id'] == row['Child']]
        return child_data['forward_week_returns'].sum()

    improvements['Total Child Return'] = improvements.apply(calculate_total_child_return, axis=1)

    # Calculate "Total Parent Return after Child Optimisation Date"
    def calculate_total_parent_return_after_child_date(row):
        # Get the optimisation_date for the Child
        child_data = build_scores[build_scores['strategy_id'] == row['Child']]
        if child_data.empty:
            return 0.0  # No data for this Child
        child_min_date = child_data['optimisation_date'].min()

        # Filter Parent data to dates on or after the Child's earliest optimisation_date
        parent_data = build_scores[build_scores['strategy_id'] == row['Parent']]
        parent_data_after_child_date = parent_data[parent_data['optimisation_date'] >= child_min_date]

        # Sum the forward_week_returns for the filtered Parent data
        return parent_data_after_child_date['forward_week_returns'].sum()

    improvements['Total Parent Return after Child Optimisation Date'] = improvements.apply(
        calculate_total_parent_return_after_child_date, axis=1
    )

    # New function to calculate "Child Weeks"
    def calculate_child_weeks(row):
        child_data = build_scores[build_scores['strategy_id'] == row['Child']]
        return len(child_data)

    # New function to calculate "Parent Weeks"
    def calculate_parent_weeks(row):
        # Get the optimisation_date for the Child
        child_data = build_scores[build_scores['strategy_id'] == row['Child']]
        if child_data.empty:
            return 0  # No data for this Child
        child_min_date = child_data['optimisation_date'].min()

        # Filter Parent data to dates on or after the Child's earliest optimisation_date
        parent_data = build_scores[build_scores['strategy_id'] == row['Parent']]
        parent_data_after_child_date = parent_data[parent_data['optimisation_date'] >= child_min_date]

        return len(parent_data_after_child_date)

    # Apply new functions to create "Child Weeks" and "Parent Weeks"
    improvements['Child Weeks'] = improvements.apply(calculate_child_weeks, axis=1)
    improvements['Parent Weeks After Child Optimisation Date'] = improvements.apply(calculate_parent_weeks, axis=1)

    # Add "Child Score" based on the `score` for the strategy at the `strategy_optimisation_date` of the Child
    def get_child_score(row):
        # Get the optimisation_date for the Child
        child_data = build_scores[build_scores['strategy_id'] == row['Child']]
        if not child_data.empty:
            child_min_date = child_data['optimisation_date'].min()
            child_score = child_data[child_data['optimisation_date'] == child_min_date]['score'].iloc[0]
            return child_score
        return None

    improvements['Child Score'] = improvements.apply(get_child_score, axis=1)

    # Add "Parent Score" based on the `score` for the Parent at the same `strategy_optimisation_date` as the Child
    def get_parent_score(row):
        # Get the optimisation_date for the Child
        child_data = build_scores[build_scores['strategy_id'] == row['Child']]
        if not child_data.empty:
            child_min_date = child_data['optimisation_date'].min()
            parent_data = build_scores[build_scores['strategy_id'] == row['Parent']]
            if not parent_data.empty:
                parent_score = parent_data[parent_data['optimisation_date'] == child_min_date]['score'].iloc[0]
                return parent_score
            return None
        return None

    improvements['Parent Score'] = improvements.apply(get_parent_score, axis=1)

    # Sort the DataFrame by the "Child" column
    improvements = improvements.sort_values(by="Child").reset_index(drop=True)

    return improvements

if __name__ == "__main__":

    use_requirements = True

    create_requirements_limits(settings.markets, settings.requirements, settings.limits, settings.market_requirements, settings.market_limits)

    # Retrieve build scores data
    build_scores = get_build_scores(use_requirements, settings.markets, settings.tawal_requirements,
                                    settings.market_requirements, settings.market_limits, settings.portfolio_name, settings.model_name,
                                    settings.indicator_count, settings.creation_date_start, settings.creation_date_end,
                                    settings.host, settings.strategies_database, settings.user, settings.password)
    
    trades = get_all_trades(settings.host, settings.strategies_database, settings.user, settings.password)
    calculate_usd_returns(build_scores, trades)

    # Create an Excel workbook
    workbook = openpyxl.Workbook()

    # Create a sheet for the overall build scores
    overall_sheet = workbook.active
    overall_sheet.title = "All Markets"
    overall_sheet.append(list(build_scores.columns))  # Write headers
    for row in build_scores.itertuples(index=False):
        overall_sheet.append(list(row._asdict().values()))

    # Calculate and save finder and improver stats
    finder_stats, improver_stats = calculate_finder_and_improver_stats(build_scores)

    # Create a sheet for Finder stats
    finder_sheet = workbook.create_sheet("Finder Stats")
    finder_sheet.append(list(finder_stats.columns))
    for row in finder_stats.itertuples(index=False):
        finder_sheet.append(list(row._asdict().values()))

    # Create a sheet for Improver stats
    improver_sheet = workbook.create_sheet("Improver Stats")
    improver_sheet.append(list(improver_stats.columns))
    for row in improver_stats.itertuples(index=False):
        improver_sheet.append(list(row._asdict().values()))

    # Calculate OOS Edge Statistics
    oos_edge_statistics = calculate_oos_edge_statistics(build_scores)
    oos_edge_sheet = workbook.create_sheet("OoS Edge")
    oos_edge_sheet.append(list(oos_edge_statistics.columns))
    for row in oos_edge_statistics.itertuples(index=False):
        oos_edge_sheet.append(list(row._asdict().values()))

    # Calculate Improvements Statistics
    improvements_statistics = calculate_improvements_statistics(build_scores)
    improvements_sheet = workbook.create_sheet("Improvements")
    improvements_sheet.append(list(improvements_statistics.columns))  # Write headers
    for row in improvements_statistics.itertuples(index=False):
        improvements_sheet.append(list(row))

    # Loop through each market in settings.markets
    for market in settings.markets:
        # Filter build scores for the current market
        market_scores = build_scores[build_scores['market'] == market]

        # Create a new sheet for the current market
        market_sheet = workbook.create_sheet(market)
        market_sheet.append(list(market_scores.columns))  # Write headers
        for row in market_scores.itertuples(index=False):
            market_sheet.append(list(row._asdict().values()))

    # Save the workbook
    workbook.save(f'{settings.write_all_path}/buildscores.xlsx')

    # Calculate stats for Finder and Improver items
    finder_stats, improver_stats = calculate_finder_and_improver_stats(build_scores)
    with pd.option_context('display.max_rows', None, 'display.max_columns', None, 'display.width', 1000):
        print("\nFinder Stats:")
        print(finder_stats)
        print("\nImprover Stats:")
        print(improver_stats)

