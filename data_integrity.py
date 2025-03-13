from datetime import datetime, timedelta
import pandas as pd
import openpyxl

from constants import BarMinutes, START_DAY_TRADING_HOUR, Severity, Reoccurring_day, DEFAULT_VOLUME
from database_health import write_activity, write_activity_integrity_batch, get_acceptable_gaps
from database_reference import get_holidays, get_bars
from utility_functions import get_current_start_week_datetime, calculate_start_trade_day
from feed_manager_first_rate_data import process_errors_first_rate_data
import settings 

variance_previous_close_to_open_variance_threshold = 0.1
variance_open_to_close_variance_threshold = 0.1
intrabar_price_variance_variance_threshold = 0.1

def combine_dataframes(bars_df, errors_df):

    if len(errors_df) == 0:
        return bars_df

    # Ensure DateTime columns are in datetime format
    bars_df['End'] = pd.to_datetime(bars_df['End'])
    errors_df['DateTime'] = pd.to_datetime(errors_df['DateTime'])
    
    # Set DateTime as the index
    bars_indexed = bars_df.set_index("End", drop=False)
    errors_indexed = errors_df.set_index("DateTime")
    
    # Use join (or pd.concat) to combine on the index
    combined_df = bars_indexed.join(errors_indexed, how="outer")
    return combined_df

def write_output(bars_df, all_errors, market, start_datetime, end_datetime, save_excel_bars, save_excel_errors, save_excel_merged):
    # Create an Excel workbook
    workbook = openpyxl.Workbook()

    unique_names = sorted({name for names in all_errors.values() for name in names})

    # Build a list of dictionaries—one per missing datetime.
    rows = []
    for dt, names in sorted(all_errors.items()):
        row = {'DateTime': dt}
        for name in unique_names:
            row[name] = name in names
        rows.append(row)
    errors_df = pd.DataFrame(rows)

    if save_excel_errors:
        errors_sheet = workbook.active
        errors_sheet.title = "Errors"
        errors_sheet.append(list(errors_df.columns))
        for row in errors_df.itertuples(index=False):
            errors_sheet.append(list(row))

    if save_excel_bars:
        # Create a sheet for the raw bars
        bars_sheet = workbook.create_sheet("Bars")
        bars_sheet.append(list(bars_df.columns))  # Write headers
        for row in bars_df.itertuples(index=False):
            bars_sheet.append(list(row._asdict().values()))

    combined_df = combine_dataframes(bars_df, errors_df)
    
    if len(errors_df) > 0 and save_excel_merged:
        combined_sheet = workbook.create_sheet("BarsWithErrors")
        combined_sheet.append(list(combined_df.columns))
        for row in combined_df.itertuples(index=False):
            combined_sheet.append(list(row))

    # Save the workbook
    filename = f'{settings.write_all_path}/data_integrity.{market}.{start_datetime.strftime("%Y.%m.%d")}.{end_datetime.strftime("%Y.%m.%d")}.xlsx'
    workbook.save(filename)
    print(f'Saved {filename}')

def calculate_expected_day_start_end(first_datetime, market_start_time, market_end_time):

    # Assumes market starts on the trading date, instead of markets like LE or Zs            
    first_datetime_trading_date = first_datetime.date()
    if first_datetime.hour < START_DAY_TRADING_HOUR:
        first_datetime_trading_date = first_datetime_trading_date - timedelta(days=1)

    actual_day_start = pd.Timestamp.combine(first_datetime_trading_date, market_start_time)        
    day_start = max(actual_day_start, first_datetime)  
    day_end = pd.Timestamp.combine(first_datetime_trading_date, market_end_time)        
    day_end = day_end + timedelta(days=1)
    return day_start, day_end, actual_day_start

def add_errors(combined_results, datetimes, name):
    for datetime in datetimes:
        if datetime not in combined_results:
            combined_results[datetime] = []
        combined_results[datetime].append(name)

def missing_bars(all_errors, name, all_bars):
    
    missing_datetimes = set(all_bars.loc[(all_bars['symbol'].isna() | (all_bars['symbol'] == '')) & (all_bars['Mask'] == True), 'End'])
    add_errors(all_errors, missing_datetimes, name)

def variance_previous_close_to_open(all_errors, name, all_bars, variance_threshold):

    previous_close = None
    exceed_threshold_datetimes = []

    for row in all_bars.itertuples():

        if previous_close is None:
            previous_close = row.close
            continue

        variance = abs(row.open - previous_close) / previous_close
        if variance > variance_threshold:
            exceed_threshold_datetimes.append(row.End)
        previous_close = row.close

    if len(exceed_threshold_datetimes) > 0:
        add_errors(all_errors, exceed_threshold_datetimes, name)

def variance_open_to_close(all_errors, name, all_bars, variance_threshold):

    exceed_threshold_datetimes = []

    for row in all_bars.itertuples():

        variance = abs(row.open - row.close) / row.close
        if variance > variance_threshold:
            exceed_threshold_datetimes.append(row.End)

    if len(exceed_threshold_datetimes) > 0:
        add_errors(all_errors, exceed_threshold_datetimes, name)

def zero_volume(all_errors, name, all_bars):

    zero_volume_datetimes = []

    for row in all_bars.itertuples():

        if row.volume == 0:
            zero_volume_datetimes.append(row.End)

    if len(zero_volume_datetimes) > 0:
        add_errors(all_errors, zero_volume_datetimes, name)

def missing_volume(all_errors, name, all_bars):

    missing_volume_datetimes = []

    for row in all_bars.itertuples():

        if row.volume == DEFAULT_VOLUME:
            missing_volume_datetimes.append(row.End)

    if len(missing_volume_datetimes) > 0:
        add_errors(all_errors, missing_volume_datetimes, name)

def intrabar_price_variance(all_errors, name, all_bars, variance_threshold):

    price_variance_datetimes = []

    for row in all_bars.itertuples():

        values = [row.open, row.high, row.low, row.close]
        exceeds_threshold = False

        for i in range(len(values)):
            for j in range(i + 1, len(values)):
                base = min(values[i], values[j])
                percentage_diff = abs(values[i] - values[j]) / base
                if percentage_diff > variance_threshold:
                    exceeds_threshold = True
                    break
            if exceeds_threshold:
                break

        if exceeds_threshold:
            price_variance_datetimes.append(row.End)

    if len(price_variance_datetimes) > 0:
        add_errors(all_errors, price_variance_datetimes, name)

def stale_price(all_errors, name, all_bars):

    stale_price_datetimes = []

    for row in all_bars.itertuples():

        if (row.open == row.high) and \
            (row.high == row.low) and \
            (row.low == row.close):
            stale_price = True
        else:
            stale_price = False

        if stale_price:
            stale_price_datetimes.append(row.End)

    if len(stale_price_datetimes) > 0:
        add_errors(all_errors, stale_price_datetimes, name)

def handle_acceptable_gaps(all_bars, acceptable_gaps):

    # Process non-reoccurring gaps
    non_reoccurring = acceptable_gaps[acceptable_gaps['reoccur_day_of_week'] == Reoccurring_day.NotReoccurring.value]
    
    for _, gap in non_reoccurring.iterrows():
        # Find all bars that fall within this gap's time range
        mask = (all_bars['Start'] >= gap['start_datetime']) & (all_bars['Start'] < gap['end_datetime'])
        
        # Update description for the matching bars
        all_bars.loc[mask, 'Description'] = gap['description']
        
        # Update mask to mark these as acceptable gaps
        all_bars.loc[mask, 'Mask'] = False
    
    # Process reoccurring gaps
    reoccurring = acceptable_gaps[acceptable_gaps['reoccur_day_of_week'] != Reoccurring_day.NotReoccurring.value]
    
    for _, gap in reoccurring.iterrows():
        day_of_week = gap['reoccur_day_of_week']
        start_time = gap['reoccur_start_time']
        end_time = gap['reoccur_end_time']
        
        # For Everyday type
        if day_of_week == Reoccurring_day.Everyday.value:           
            # Find all bars within the time range for any day            
            mask = (all_bars['Start'] >= gap['start_datetime']) & \
                   (all_bars['Start'] < gap['end_datetime']) & \
                   (all_bars['Start'].dt.time >= start_time) & \
                   (all_bars['Start'].dt.time < end_time)
            
        # For specific days of the week
        else:
            # Convert enum value to corresponding day of week (Python's weekday is 0-6, Monday is 0)
            # Adjust from our enum where Monday is 2
            python_weekday = day_of_week - 2 if day_of_week > 1 else 6  # Sunday would be 6 in Python's scheme     
            
            # Find all bars on the specified day of week and within the time range
            mask = (all_bars['Start'] >= gap['start_datetime']) & \
                   (all_bars['Start'] < gap['end_datetime']) & \
                   (all_bars['Start'].dt.weekday == python_weekday) & \
                   (all_bars['Start'].dt.time >= start_time) & \
                   (all_bars['Start'].dt.time < end_time)
        
        # Update description and mask for the matching bars
        all_bars.loc[mask, 'Description'] = gap['description']
        all_bars.loc[mask, 'Mask'] = False

def calculate_tradable_array(market, start_datetime, end_datetime,
                             host, strategies_database, user, password): 

    full_time_range = pd.date_range(start=start_datetime, end=end_datetime, freq='1 min')
    end_times = full_time_range + pd.Timedelta(minutes=1)
    
    # Create DataFrame directly from the time range
    all_bars = pd.DataFrame({
        # Exclude the last start time since we don't have its end
        'Start': full_time_range[:-1],  
        'End': end_times[:-1]        
    })
    
    # Add remaining columns with default values
    all_bars['Mask'] = True
    all_bars['Description'] = ''

    acceptable_gaps = get_acceptable_gaps(market, host, strategies_database, user, password)   
    handle_acceptable_gaps(all_bars, acceptable_gaps)

    holidays_df = get_holidays(host, strategies_database, user, password)
    handle_holidays(all_bars, holidays_df)
    
    return all_bars

def handle_holidays(all_bars, holidays_df):
    # Check if holidays_df exists and is not empty
    if holidays_df is None or holidays_df.empty:
        return all_bars
    
    # Iterate through each holiday
    for _, holiday in holidays_df.iterrows():
        start_time = holiday['start']
        end_time = holiday['end']
        holiday_name = holiday['name']
        
        # Find all bars that fall within this holiday's time range
        mask = (all_bars['Start'] >= start_time) & (all_bars['Start'] < end_time)
        
        # Update description for the matching bars
        all_bars.loc[mask, 'Description'] = holiday_name
        
        # Update mask to mark these as non-tradable periods
        all_bars.loc[mask, 'Mask'] = False
    
    return all_bars

def combine_bars(all_bars, historical_bars):

    if historical_bars is None or historical_bars.empty:
        return

    all_bars_indexed = all_bars.set_index("End", drop=False)
    historical_bars_indexed = historical_bars.set_index("datetime")
    
    # Use join (or pd.concat) to combine on the index
    joined_bars = all_bars_indexed.join(historical_bars_indexed, how="outer")    

    return joined_bars
       
def check_data_integrity(market, start_datetime, end_datetime, given_bars = None, 
                         save_excel = False, save_activity = False,
                         save_excel_bars = False, save_excel_errors = False, save_excel_merged = False):

    if save_activity:
        write_activity('DataIntegrity', datetime.now(), Severity.Info, f'Started data integrity checks on {market} between {start_datetime} and {end_datetime}', 
                       False, save_activity, settings.host, settings.strategies_database, settings.user, settings.password)
        
    all_errors = {}
        
    all_bars = calculate_tradable_array(market, start_datetime, end_datetime, 
                                        settings.host, settings.strategies_database, settings.user, settings.password)
    
    bars_database = given_bars
    if given_bars is None:    
        bars_database = get_bars(settings.host, settings.user, settings.password, market, start_datetime, end_datetime, BarMinutes.Minute1.value)

    if bars_database is None:
        write_activity('DataIntegrity', datetime.now(), Severity.Error, f'Error loading bars for data integrity checks on {market} between {start_datetime} and {end_datetime}', 
                       False, save_activity, settings.host, settings.strategies_database, settings.user, settings.password)
        all_errors[start_datetime] = ['BarLoadError']
        return all_errors

    all_bars = combine_bars(all_bars, bars_database)

    missing_bars(all_errors, 'missing_bars', all_bars)
    zero_volume(all_errors, 'zero_volume', all_bars)
    missing_volume(all_errors, 'missing_volume', all_bars)
    stale_price(all_errors, 'stale_price', all_bars)
    variance_previous_close_to_open(all_errors, 'variance_previous_close_to_open', 
                                    all_bars, variance_previous_close_to_open_variance_threshold)
    variance_open_to_close(all_errors, 'variance_open_to_close', 
                           all_bars, variance_open_to_close_variance_threshold)
    intrabar_price_variance(all_errors, 'intrabar_price_variance', 
                            all_bars, intrabar_price_variance_variance_threshold)

    if save_excel and all_bars is not None and len(all_bars) > 0:
        write_output(all_bars, all_errors, market, start_datetime, end_datetime, save_excel_bars, save_excel_errors, save_excel_merged)

    if save_activity:
        write_activity_integrity_batch(all_errors, datetime.now(), market, "DataIntegrity", Severity.Error, save_activity,
                                       settings.host, settings.strategies_database, settings.user, settings.password)

        write_activity('DataIntegrity', datetime.now(), Severity.Info, f'Finished data integrity checks on {market} between {start_datetime} and {end_datetime}', 
                       False, save_activity, settings.host, settings.strategies_database, settings.user, settings.password)

    return all_errors

def check_data_integrity_pass_dates(market, start_datetime, end_datetime, save_excel = False, save_activity = False,
                              save_excel_bars = False, save_excel_errors = False, save_excel_merged = False):
    
    combined_results = check_data_integrity(market, start_datetime, end_datetime, save_excel, save_activity,
                                            save_excel_bars, save_excel_errors, save_excel_merged)
    return len(combined_results) == 0

def check_data_integrity_pass_current_week(market, save_excel = False, save_activity = False,
                              save_excel_bars = False, save_excel_errors = False, save_excel_merged = False):
    
    start_week = get_current_start_week_datetime()
    next_week = start_week + timedelta(weeks=1)

    combined_results = check_data_integrity(market, start_week, next_week, save_excel, save_activity,
                                            save_excel_bars, save_excel_errors, save_excel_merged)
    return len(combined_results) == 0

def check_data_integrity_pass_current_day(market, save_excel = False, save_activity = False,
                                          save_excel_bars = False, save_excel_errors = False, save_excel_merged = False):
    
    start_day = calculate_start_trade_day(datetime.now())
    next_day = start_day + timedelta(days=1)

    combined_results = check_data_integrity(market, start_day, next_day, save_excel, save_activity,
                                            save_excel_bars, save_excel_errors, save_excel_merged)
    return len(combined_results) == 0

def check_data_integrity_pass_all_time(market, save_excel = False, save_activity = False,
                                       save_excel_bars = False, save_excel_errors = False, save_excel_merged = False):
    
    start_datetime = settings.start_pd
    end_datetime = datetime.now()    

    combined_results = check_data_integrity(market, start_datetime, end_datetime, save_excel, save_activity,
                                            save_excel_bars, save_excel_errors, save_excel_merged)
    return len(combined_results) == 0


if __name__ == "__main__":

    market = 'NQ'
    start_datetime = '2008-01-06 17:00'
    # end_datetime = '2025-03-02 17:00'
    # end_datetime = '2008-03-13 17:00'
    end_datetime = '2009-01-06 17:00'

    save_excel = True
    save_excel_bars = True
    save_excel_errors = True
    save_excel_merged = True
    save_activity = False
    run_errors_on_first_rate_date = False

    start_pd = pd.Timestamp(start_datetime)
    end_pd = pd.Timestamp(end_datetime)

    # check_data_integrity_pass_current_week(market, save_excel, save_activity, 
    #                                        save_excel_bars, save_excel_errors, save_excel_merged)

    error_bars = check_data_integrity(market, start_pd, end_pd, save_excel, save_activity, 
                                      save_excel_bars, save_excel_errors, save_excel_merged)
    
    if len(error_bars) > 0 and run_errors_on_first_rate_date:
        process_errors_first_rate_data(market, error_bars)
